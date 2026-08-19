# -*- coding: utf-8 -*-
"""
Risk-Based EVM Analyzer
Alat analisis kinerja biaya & waktu proyek EPC berbasis risiko - mengintegrasikan
AHP (pembobotan risiko), Earned Value Management, dan simulasi Monte Carlo
(Augmented EVM, mengacu Duc 2025) ke dalam satu dashboard interaktif.

Cara menjalankan:
    pip install streamlit pandas numpy matplotlib openpyxl plotly
    streamlit run dashboard_pro.py

Catatan: seluruh logika perhitungan direplikasi persis dari skrip yang sudah
divalidasi (augmented_evm_v3_boq_riskfactor.py) - file ini murni lapisan
tampilan + narasi interpretasi, tidak ada perhitungan baru yang berbeda.
"""

import io
import os
import base64
from datetime import datetime
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font as XLFont, PatternFill as XLFill, Alignment as XLAlign, Border as XLBorder, Side as XLSide
from openpyxl.worksheet.datavalidation import DataValidation
import streamlit as st
import plotly.graph_objects as go
import warnings
warnings.filterwarnings("ignore")

# =========================================================================
# IDENTITAS INSTANSI & PENYUSUN
# =========================================================================
UNIVERSITAS = "Universitas Islam Indonesia"
FAKULTAS = "Fakultas Teknik Desain Inovasi"
JURUSAN = "Jurusan Teknik Sipil"
PRODI = "Program Studi Teknik Sipil (Program Sarjana)"
PENYUSUN_NAMA = "Dea Aprilia Indrawati"
PENYUSUN_NIM = "22511032"
LOGO_PATH = os.path.join(os.path.dirname(__file__), "UII - Biru.png")

# Brand alat/dashboard itu sendiri (terpisah dari identitas kampus) - PRIME:
# Professional Resource for Infrastructure Management & Engineering.
PRIME_NAME = "PRIME"
PRIME_TAGLINE = "Professional Resource for Infrastructure Management & Engineering"
PRIME_LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo prime.png")

# Studi kasus yang sedang dimuat berjalan pada proyek milik klien berikut -
# ditampilkan sbg chip identitas terpisah dari brand alat/kampus.
KLIEN_PROYEK = "PT Pertamina Patra Niaga"

@st.cache_data(show_spinner=False)
def _load_logo_b64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except FileNotFoundError:
        return None

LOGO_B64 = _load_logo_b64(LOGO_PATH)
PRIME_LOGO_B64 = _load_logo_b64(PRIME_LOGO_PATH)

def _brand_ribbon_html():
    """Strip identitas paling atas: brand alat (PRIME) di kiri, identitas
    kampus + chip klien studi kasus di kanan. Dipakai di landing page
    maupun halaman dashboard utama supaya konsisten."""
    prime_img = f'<img src="data:image/png;base64,{PRIME_LOGO_B64}" alt="Logo {PRIME_NAME}" />' if PRIME_LOGO_B64 else ""
    return f"""
    <div class="brand-ribbon">
      <div class="brand-left">
        {prime_img}
        <div>
          <div class="brand-name">{PRIME_NAME}</div>
          <div class="brand-tagline">{PRIME_TAGLINE}</div>
        </div>
      </div>
      <div class="brand-right">
        <span class="client-chip"><span class="dot2"></span>Studi Kasus: {KLIEN_PROYEK}</span>
        <div class="inst-block">
          <div class="inst-univ">{UNIVERSITAS}</div>
          <div class="inst-fac">{FAKULTAS} — {JURUSAN}</div>
        </div>
      </div>
    </div>
    """

# =========================================================================
# KONFIGURASI HALAMAN & PALET WARNA
# =========================================================================
st.set_page_config(
    page_title="Risk-Based EVM Analyzer",
    page_icon=PRIME_LOGO_PATH if os.path.exists(PRIME_LOGO_PATH) else None,
    layout="wide",
    initial_sidebar_state="expanded",
)

# Palet korporat: navy sebagai warna identitas utama, netral abu-biru untuk
# latar & teks, dan tiga warna status standar (baik/waspada/kritis) yang
# konsisten dipakai di seluruh badge, KPI, dan grafik.
C = {
    "navy": "#0B2942", "navy2": "#123A5C", "accent": "#1E6FA6", "accent_lt": "#D8E8F2",
    "good": "#1B7A43", "good_bg": "#E3F3E8", "bad": "#B3261E", "bad_bg": "#FBE7E5",
    "warn": "#B8720C", "warn_bg": "#FCF0D9",
    "bg": "#F2F4F7", "card": "#FFFFFF", "ink": "#1A2733", "ink_dim": "#5B6B79",
    "border": "#DFE3E8",
    # alias dipertahankan agar kode lama yang mereferensikan kunci ini tetap jalan
    "amber": "#1E6FA6", "amber_lt": "#D8E8F2", "teal": "#1B7A43", "teal_lt": "#E3F3E8",
    "red": "#B3261E", "red_lt": "#FBE7E5", "gold": "#B8720C",
    # Warna identitas tambahan - diambil langsung dari dua situs referensi:
    # emas FTDI UII (ftdi.uii.ac.id, header navy+emas akademik) dan
    # merah/biru Pertamina (pertamina.com, brand korporat klien studi kasus).
    "fac_navy": "#06337B", "fac_gold": "#F7D217",
    "client_red": "#E21F23", "client_blue": "#0B2F9F",
    "prime_navy": "#26597C",
}

RI_TABLE = {1: 0, 2: 0, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45}
_BOQ_FALLBACK = {
    "I":   {"material": 0.0,                 "jasa": 1248439423.508313},
    "II":  {"material": 1046607274.144194,    "jasa": 509832863.1320728},
    "III": {"material": 350761098.40007687,   "jasa": 316123890.8866162},
    "IV":  {"material": 1540000.0000000002,   "jasa": 132411512.02180462},
    "V":   {"material": 0.0,                  "jasa": 164859402.30245006},
}

# =========================================================================
# TEMPLATE EXCEL UNDUH-MANDIRI
# Supaya dashboard ini benar-benar bisa dipakai pekerja/proyek lain (bukan
# hanya studi kasus TA ini), disediakan template kosong yang strukturnya
# PERSIS mengikuti apa yang dibaca fungsi load_and_compute() di atas - isi,
# unduh, isi data proyek sendiri, unggah kembali di panel kiri.
# =========================================================================
_XL_HEAD_FILL = XLFill("solid", fgColor="0B2942")
_XL_HEAD_FONT = XLFont(bold=True, color="FFFFFF", size=10)
_XL_NOTE_FONT = XLFont(italic=True, color="B3261E", size=9)
_XL_LABEL_FONT = XLFont(bold=True, color="0B2942", size=10)
_XL_BORDER = XLBorder(*(XLSide(style="thin", color="DFE3E8"),) * 4)


@st.cache_data(show_spinner=False)
def build_evm_template_bytes():
    wb = openpyxl.Workbook()

    # --- Sheet "Data Umum" (dibaca TANPA header, posisi baris harus persis) ---
    ws = wb.active
    ws.title = "Data Umum"
    ws["A1"] = "DATA UMUM PROYEK - jangan ubah/sisipkan baris, isi kolom B saja"
    ws["A1"].font = _XL_NOTE_FONT
    ws["A2"] = "(baris ini sengaja dikosongkan - jangan dihapus)"
    rows_umum = [
        (3, "Nama Proyek", "Contoh: Proyek Pembangunan Gedung X"),
        (4, "(baris ini sengaja dikosongkan - jangan dihapus)", None),
        (5, "BAC - Budget at Completion (Rp)", 10000000000),
        (6, "Tanggal Mulai Proyek", datetime(2025, 1, 1)),
        (7, "Rencana Total Hari - Baseline AWAL", 240),
        (8, "Rencana Total Hari - Baseline REVISI (isi sama dgn baris 7 jika tidak ada amandemen)", 240),
    ]
    for r, label, val in rows_umum:
        ws.cell(row=r, column=1, value=label).font = _XL_LABEL_FONT if val is not None else _XL_NOTE_FONT
        if val is not None:
            c = ws.cell(row=r, column=2, value=val)
            if r == 6:
                c.number_format = "yyyy-mm-dd"
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 26

    # --- Sheet "Data 14 Periode" (header baris 1, nama kolom harus persis) ---
    ws2 = wb.create_sheet("Data 14 Periode")
    headers = ["Laporan Ke-", "Tanggal Cutoff", "Rencana Kumulatif (%)", "Aktual Kumulatif (%)",
               "AC Kumulatif (Rp)", "Baseline Berlaku"]
    for j, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = _XL_HEAD_FONT
        c.fill = _XL_HEAD_FILL
        c.alignment = XLAlign(horizontal="center", wrap_text=True)
        c.border = _XL_BORDER
    dv = DataValidation(type="list", formula1='"Awal,Revisi"', allow_blank=False,
                         errorTitle="Pilihan tidak valid", error="Pilih 'Awal' atau 'Revisi' dari daftar.")
    ws2.add_data_validation(dv)
    for i in range(14):
        r = i + 2
        ws2.cell(row=r, column=1, value=i + 1)
        ws2.cell(row=r, column=2).number_format = "yyyy-mm-dd"
        ws2.cell(row=r, column=3).number_format = "0.00%"
        ws2.cell(row=r, column=4).number_format = "0.00%"
        ws2.cell(row=r, column=5).number_format = "#,##0"
        dv.add(ws2.cell(row=r, column=6))
        for c in range(1, 7):
            ws2.cell(row=r, column=c).border = _XL_BORDER
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 16
    for col in "CDE":
        ws2.column_dimensions[col].width = 20
    ws2.column_dimensions["F"].width = 16
    ws2.freeze_panes = "A2"

    # --- Sheet "BOQ" (kolom B = kode seksi/baris "JUMLAH <kode>", kolom R/S = Material/Jasa) ---
    ws3 = wb.create_sheet("BOQ")
    ws3["A1"] = "Isi rincian item BoQ Anda sendiri di baris-baris kosong tiap seksi (opsional, bebas format)."
    ws3["A1"].font = _XL_NOTE_FONT
    ws3["A2"] = "WAJIB diisi: baris 'JUMLAH <kode>' pada tiap seksi - kolom R (Material) & S (Jasa)."
    ws3["A2"].font = _XL_NOTE_FONT
    for j, h in [(2, "Kode Seksi"), (3, "Uraian Pekerjaan"), (18, "Material (Rp)"), (19, "Jasa (Rp)")]:
        c = ws3.cell(row=4, column=j, value=h)
        c.font = _XL_HEAD_FONT
        c.fill = _XL_HEAD_FILL
    seksi_names = {"I": "Persiapan", "II": "Pekerjaan Utama", "III": "Pekerjaan Sekunder",
                   "IV": "Pengetesan/Commissioning", "V": "Finishing"}
    r = 5
    for kode, nama in seksi_names.items():
        ws3.cell(row=r, column=2, value=kode).font = _XL_LABEL_FONT
        ws3.cell(row=r, column=3, value=f"SEKSI {kode} - {nama}").font = _XL_LABEL_FONT
        r += 1
        ws3.cell(row=r, column=3, value="(sisipkan baris item pekerjaan di sini jika perlu)").font = _XL_NOTE_FONT
        r += 1
        ws3.cell(row=r, column=2, value=f"JUMLAH {kode}").font = _XL_LABEL_FONT
        ws3.cell(row=r, column=18, value=0).number_format = "#,##0"
        ws3.cell(row=r, column=19, value=0).number_format = "#,##0"
        r += 2
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 34
    ws3.column_dimensions["R"].width = 18
    ws3.column_dimensions["S"].width = 18

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@st.cache_data(show_spinner=False)
def build_ahp_template_bytes():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    kriteria = {
        "Kategori": ["A - Perancangan", "B - Pengadaan", "C - Konstruksi"],
        "Subfaktor A": ["A1 - Kesadaran efisiensi biaya perancangan", "A2 - Kompetensi personel perancangan", "A3 - Komunikasi & koordinasi"],
        "Subfaktor B": ["B1 - Pemilihan/pengelolaan pemasok", "B2 - Pemilihan/pengelolaan subkontraktor", "B3 - Pengendalian biaya pengadaan"],
        "Subfaktor C": ["C1 - Kompetensi personel manajemen lapangan", "C2 - Pengendalian mutu, harga, termin", "C3 - Metode pengelolaan biaya konstruksi"],
    }
    responden_label = ["Responden 1 (mis. Project Manager)", "Responden 2 (mis. Site Engineer)", "Responden 3 (mis. Project Control)"]

    ws0 = wb.create_sheet("Petunjuk")
    ws0["A1"] = "PETUNJUK PENGISIAN MATRIKS PERBANDINGAN BERPASANGAN (AHP - Skala Saaty)"
    ws0["A1"].font = XLFont(bold=True, size=12, color="0B2942")
    petunjuk_lines = [
        "",
        "Setiap sheet (Kategori, Subfaktor A, Subfaktor B, Subfaktor C) berisi 3 matriks 3x3,",
        "satu untuk tiap responden ahli proyek (Project Manager, Site Engineer, Project Control",
        "atau peran setara di proyek Anda).",
        "",
        "Cara mengisi tiap sel (baris i dibandingkan kolom j):",
        "  1 = sama penting        3 = sedikit lebih penting     5 = lebih penting",
        "  7 = sangat lebih penting   9 = mutlak lebih penting   (2,4,6,8 = nilai antara)",
        "  Jika kolom j lebih penting dari baris i, isi kebalikannya, mis. =1/3, =1/5, dst.",
        "",
        "Diagonal (baris i = kolom j) WAJIB bernilai 1 - jangan diubah.",
        "Sel di bawah diagonal harus kebalikan matematis dari sel di atas diagonal",
        "(mis. jika [1,2]=3 maka [2,1] harus =1/3) - konsisten dengan Persamaan 3.2 Bab III.",
        "",
        "Setelah semua sheet terisi, unggah file ini di panel kiri dashboard bersama file data EVM.",
    ]
    for i, line in enumerate(petunjuk_lines, start=2):
        ws0.cell(row=i, column=1, value=line)
    ws0.column_dimensions["A"].width = 90

    for sheet_name, labels in kriteria.items():
        ws = wb.create_sheet(sheet_name)
        ws.column_dimensions["A"].width = 4
        for j in range(3):
            ws.column_dimensions[chr(ord("B") + j)].width = 16
        for block_i, r0 in enumerate([6, 12, 18]):
            ws.cell(row=r0 - 1, column=1, value=responden_label[block_i]).font = _XL_LABEL_FONT
            for j, lab in enumerate(labels):
                c = ws.cell(row=r0 - 1, column=2 + j, value=lab)
                c.font = XLFont(bold=True, size=8, color="FFFFFF")
                c.fill = _XL_HEAD_FILL
                c.alignment = XLAlign(wrap_text=True, horizontal="center")
            for i in range(3):
                for j in range(3):
                    cell = ws.cell(row=r0 + i, column=2 + j)
                    cell.value = 1
                    cell.border = _XL_BORDER
                    cell.alignment = XLAlign(horizontal="center")
                    if i == j:
                        cell.fill = XLFill("solid", fgColor="E3F3E8")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# =========================================================================
# CSS KUSTOM - TAMPILAN KORPORAT (gaya dashboard instansi/BUMN)
# =========================================================================
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] {{ font-family: 'Inter', sans-serif; }}
.stApp {{ background-color: {C['bg']}; }}

/* Sembunyikan chrome bawaan Streamlit yang tidak perlu - TAPI jangan
   sembunyikan <header> secara utuh, karena tombol buka/tutup sidebar
   menempel di situ pada versi Streamlit terbaru. Menyembunyikannya
   membuat sidebar tidak bisa dibuka lagi setelah tertutup/tampilan
   diperbesar. Sembunyikan isi toolbar-nya saja (menu, tombol Deploy,
   ikon GitHub), dan pastikan kontrol sidebar selalu di atas & terlihat. */
#MainMenu {{visibility: hidden;}}
footer {{visibility: hidden;}}
header[data-testid="stHeader"] {{
    background: transparent; height: 2.75rem;
}}
header[data-testid="stHeader"] [data-testid="stToolbar"] {{visibility: hidden;}}
header[data-testid="stHeader"] [data-testid="stDecoration"] {{display: none;}}
/* PENTING: visibility:hidden pada stToolbar di atas TURUN ke elemen anak
   (visibility, tidak seperti display, diwariskan) - termasuk tombol buka
   sidebar (stExpandSidebarButton) yang ternyata bersarang di dalam
   stToolbar pada versi Streamlit ini. Ini akar penyebab sidebar tidak
   bisa dibuka lagi setelah ditutup. Kembalikan visibility eksplisit utk
   tombol ini & semua isinya (ikon di dalamnya jg mewarisi hidden). */
[data-testid="stExpandSidebarButton"],
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapseButton"] {{
    visibility: visible !important; display: flex !important;
    opacity: 1 !important; pointer-events: auto !important;
    z-index: 999999 !important; position: relative;
}}
[data-testid="stExpandSidebarButton"] *,
[data-testid="stSidebarCollapsedControl"] *,
[data-testid="collapsedControl"] *,
[data-testid="stSidebarCollapseButton"] * {{
    visibility: visible !important;
    opacity: 1 !important; pointer-events: auto !important;
}}

/* Strip identitas instansi di bagian paling atas */
.topbar {{
    display: flex; align-items: center; justify-content: space-between;
    background: {C['navy']}; color: #C9D6E3; padding: 0.5rem 1.4rem;
    border-radius: 0; font-size: 0.76rem; letter-spacing: 0.04em;
    text-transform: uppercase; font-weight: 600;
}}
.topbar span.dot {{ color: {C['accent']}; margin-right: 0.4rem; }}

/* Ribbon identitas gabungan: brand alat (PRIME) + instansi akademik (gaya
   header navy/emas FTDI UII) + chip klien studi kasus (gaya merah/biru
   Pertamina) - satu strip HERO di paling atas halaman, dibuat full-bleed
   (menembus padding container Streamlit) supaya benar-benar penuh selebar
   layar, bukan cuma sebatas lebar konten. */
div[data-testid="stMainBlockContainer"] {{ padding-top: 0.5rem !important; }}
.brand-ribbon {{
    display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap;
    gap: 0.9rem; background: linear-gradient(120deg, {C['fac_navy']} 0%, {C['prime_navy']} 100%);
    padding: 0.7rem 2vw; border-bottom: 4px solid {C['fac_gold']};
    width: 100vw; position: relative; left: 50%; right: 50%;
    margin-left: -50vw; margin-right: -50vw; margin-top: 0; margin-bottom: 1.2rem;
    box-shadow: 0 2px 10px rgba(6,51,123,0.2); z-index: 1;
}}
.brand-ribbon .brand-left {{ display: flex; align-items: center; gap: 0.7rem; }}
.brand-ribbon .brand-left img {{ height: 34px; width: auto; filter: drop-shadow(0 1px 3px rgba(0,0,0,0.25)); }}
.brand-ribbon .brand-name {{ font-size: 1.2rem; font-weight: 800; color: #FFFFFF; letter-spacing: 0.02em; line-height: 1.1; }}
.brand-ribbon .brand-tagline {{ font-size: 0.68rem; color: {C['fac_gold']}; font-weight: 600; letter-spacing: 0.01em; margin-top: 0.1rem; }}
.brand-ribbon .brand-right {{ display: flex; align-items: center; gap: 0.8rem; flex-wrap: wrap; justify-content: flex-end; }}
.brand-ribbon .inst-block {{ text-align: right; line-height: 1.3; }}
.brand-ribbon .inst-block .inst-univ {{ font-size: 0.78rem; font-weight: 700; color: #FFFFFF; }}
.brand-ribbon .inst-block .inst-fac {{ font-size: 0.66rem; color: {C['fac_gold']}; font-weight: 600; }}
.client-chip {{
    display: inline-flex; align-items: center; gap: 0.35rem; background: {C['client_red']};
    color: #FFFFFF; font-size: 0.68rem; font-weight: 700; letter-spacing: 0.02em;
    padding: 0.3rem 0.7rem; border-radius: 99px; text-transform: uppercase;
    box-shadow: 0 2px 6px rgba(226,31,35,0.35);
}}
.client-chip .dot2 {{ width: 7px; height: 7px; border-radius: 50%; background: {C['client_blue']}; display: inline-block; }}
@media (max-width: 900px) {{
  .brand-ribbon {{ flex-direction: column; align-items: flex-start; }}
  .brand-ribbon .brand-right {{ justify-content: flex-start; }}
  .brand-ribbon .inst-block {{ text-align: left; }}
}}

/* Panel identitas proyek (pengganti hero banner) */
.page-head {{
    background: {C['card']}; border: 1px solid {C['border']}; border-left: 5px solid {C['navy']};
    border-radius: 0 0 10px 10px; padding: 1.4rem 1.6rem 1.5rem 1.5rem;
    margin-bottom: 1.6rem;
}}
.page-head .eyebrow {{
    font-size: 0.72rem; letter-spacing: 0.12em; text-transform: uppercase;
    color: {C['ink_dim']}; font-weight: 700; margin-bottom: 0.35rem;
}}
.page-head h1 {{
    font-family: 'Inter', sans-serif; font-size: 1.55rem; font-weight: 800;
    margin: 0 0 0.55rem 0; color: {C['navy']};
}}
.page-head p {{ font-size: 0.92rem; color: {C['ink_dim']}; max-width: 90ch; margin: 0; line-height: 1.55; }}

/* Kartu metrik / KPI */
.metric-card {{
    background: {C['card']}; border: 1px solid {C['border']}; border-radius: 10px;
    padding: 1.05rem 1.25rem; height: 100%; border-top: 3px solid {C['navy']};
}}
.metric-label {{
    font-size: 0.7rem; letter-spacing: 0.05em; text-transform: uppercase;
    color: {C['ink_dim']}; font-weight: 700; margin-bottom: 0.4rem;
}}
.metric-value {{ font-size: 1.5rem; font-weight: 800; color: {C['navy']}; line-height: 1.15; }}
.metric-sub {{ font-size: 0.78rem; margin-top: 0.35rem; font-weight: 600; }}
.metric-bar-track {{ background: {C['bg']}; border-radius: 99px; height: 6px; margin-top: 0.6rem; overflow: hidden; }}
.metric-bar-fill {{ height: 100%; border-radius: 99px; }}

/* Badge status */
.badge {{
    display: inline-block; padding: 0.26rem 0.8rem; border-radius: 4px;
    font-size: 0.75rem; font-weight: 700; letter-spacing: 0.02em;
}}
.badge-good {{ background: {C['good_bg']}; color: {C['good']}; }}
.badge-bad {{ background: {C['bad_bg']}; color: {C['bad']}; }}
.badge-warn {{ background: {C['warn_bg']}; color: {C['warn']}; }}

/* Kartu catatan analisis (pengganti "insight box") */
.insight {{
    background: {C['card']}; border: 1px solid {C['border']}; border-left: 4px solid {C['accent']};
    border-radius: 6px; padding: 0.95rem 1.25rem; margin: 0.85rem 0 1.2rem 0;
    font-size: 0.9rem; line-height: 1.62; color: {C['ink']};
}}
.insight b {{ color: {C['navy']}; }}
.insight-title {{
    font-size: 0.7rem; letter-spacing: 0.08em; text-transform: uppercase;
    color: {C['accent']}; font-weight: 700; margin-bottom: 0.4rem;
}}

/* Section heading */
.section-h {{
    font-family: 'Inter', sans-serif; font-size: 1.05rem; font-weight: 800;
    text-transform: uppercase; letter-spacing: 0.03em;
    color: {C['navy']}; margin: 1.7rem 0 0.3rem 0; padding-bottom: 0.55rem;
    border-bottom: 2px solid {C['navy']};
}}
.section-sub {{ color: {C['ink_dim']}; font-size: 0.86rem; margin-bottom: 1rem; }}

/* Step card di landing page */
.step-card {{
    background: {C['card']}; border: 1px solid {C['border']}; border-radius: 10px;
    padding: 1.25rem; text-align: left; height: 100%; border-top: 3px solid {C['accent']};
}}
.step-num {{
    display: inline-flex; align-items: center; justify-content: center;
    width: 26px; height: 26px; border-radius: 4px; background: {C['navy']};
    color: white; font-weight: 800; font-size: 0.8rem; margin-bottom: 0.7rem;
}}
.step-card h4 {{ margin: 0.3rem 0; color: {C['navy']}; font-size: 0.95rem; font-weight: 700; }}
.step-card p {{ color: {C['ink_dim']}; font-size: 0.83rem; line-height: 1.5; margin: 0; }}

/* Tabel & dataframe: header tegas, garis rapi */
[data-testid="stDataFrame"] {{ border: 1px solid {C['border']}; border-radius: 8px; overflow: hidden; }}

/* Tab navigasi */
.stTabs [data-baseweb="tab-list"] {{ gap: 0; border-bottom: 2px solid {C['border']}; }}
.stTabs [data-baseweb="tab"] {{
    font-weight: 600; font-size: 0.88rem; color: {C['ink_dim']}; padding: 0.7rem 1.1rem;
}}
.stTabs [aria-selected="true"] {{ color: {C['navy']} !important; }}

/* Sidebar navigasi */
[data-testid="stSidebar"] {{ background-color: {C['navy']}; }}
[data-testid="stSidebar"] * {{ color: #DCE5EE !important; }}
[data-testid="stSidebar"] .sidebar-brand {{
    font-size: 0.95rem; font-weight: 800; color: #FFFFFF !important; letter-spacing: 0.01em;
}}
[data-testid="stSidebar"] .sidebar-caption {{ font-size: 0.78rem; color: #A9BCCE !important; line-height: 1.5; }}
[data-testid="stSidebar"] .sidebar-step {{
    font-size: 0.7rem; letter-spacing: 0.08em; text-transform: uppercase; font-weight: 700;
    color: {C['accent']} !important; margin-top: 1.1rem; margin-bottom: 0.3rem;
}}
[data-testid="stSidebar"] hr {{ border-color: #23456A; }}

/* Identitas instansi di sidebar */
.sidebar-institution {{
    display: flex; align-items: center; gap: 0.7rem; margin-bottom: 1rem;
    padding-bottom: 1rem; border-bottom: 1px solid #23456A;
}}
.sidebar-institution img {{ width: 42px; height: auto; flex-shrink: 0; }}
.sidebar-institution .inst-name {{ font-size: 0.72rem; font-weight: 700; color: #FFFFFF !important; line-height: 1.3; }}
.sidebar-institution .inst-faculty {{ font-size: 0.68rem; color: #A9BCCE !important; line-height: 1.3; margin-top: 0.15rem; }}
.sidebar-institution .inst-jurusan {{ font-size: 0.66rem; color: {C['fac_gold']} !important; line-height: 1.3; font-weight: 600; }}
.sidebar-prime {{ display: flex; align-items: center; gap: 0.55rem; margin-bottom: 0.9rem; }}
.sidebar-prime img {{ height: 26px; width: auto; }}
.sidebar-prime .prime-word {{ font-size: 0.9rem; font-weight: 800; color: #FFFFFF !important; letter-spacing: 0.02em; }}
.sidebar-prime .prime-tag {{ font-size: 0.6rem; color: {C['fac_gold']} !important; font-weight: 600; }}

/* Profil penyusun di footer halaman */
.author-card {{
    display: flex; align-items: center; gap: 0.9rem; margin-top: 1.5rem;
    padding-top: 1.2rem; border-top: 1px solid {C['border']};
}}
.author-card img {{ width: 34px; height: auto; flex-shrink: 0; }}
.author-card .author-label {{ font-size: 0.68rem; letter-spacing: 0.06em; text-transform: uppercase; color: {C['ink_dim']}; font-weight: 700; }}
.author-card .author-name {{ font-size: 0.86rem; font-weight: 700; color: {C['navy']}; }}
.author-card .author-nim {{ font-size: 0.78rem; color: {C['ink_dim']}; }}
</style>
""", unsafe_allow_html=True)

PLOT_TEMPLATE = "plotly_white"
PLOT_COLORWAY = [C["navy"], C["accent"], C["good"], C["warn"], C["bad"], C["navy2"]]
def style_fig(fig, height=380, title=None):
    layout_kwargs = dict(
        template=PLOT_TEMPLATE, height=height, margin=dict(l=10, r=10, t=45 if title else 15, b=10),
        font=dict(family="Inter, sans-serif", size=12, color=C["ink"]),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
        plot_bgcolor="white", paper_bgcolor="white",
        hovermode="x unified", colorway=PLOT_COLORWAY,
    )
    if title:
        layout_kwargs["title"] = dict(text=title, font=dict(size=13, color=C["navy"]))
    fig.update_layout(**layout_kwargs)
    fig.update_xaxes(gridcolor="#EAEDF1", showline=True, linecolor=C["border"])
    fig.update_yaxes(gridcolor="#EAEDF1", showline=True, linecolor=C["border"])
    return fig

def metric_card(label, value, sub=None, sub_color=None, pct=None, pct_color=None):
    sub_html = f'<div class="metric-sub" style="color:{sub_color or C["ink_dim"]}">{sub}</div>' if sub else ""
    bar_html = ""
    if pct is not None:
        pct_clamped = max(0, min(100, pct))
        bar_html = (f'<div class="metric-bar-track"><div class="metric-bar-fill" '
                    f'style="width:{pct_clamped:.0f}%;background:{pct_color or C["accent"]}"></div></div>')
    st.markdown(f"""
    <div class="metric-card">
      <div class="metric-label">{label}</div>
      <div class="metric-value">{value}</div>
      {sub_html}
      {bar_html}
    </div>
    """, unsafe_allow_html=True)

def insight_box(title, html_body):
    st.markdown(f"""
    <div class="insight">
      <div class="insight-title">Catatan Analisis — {title}</div>
      {html_body}
    </div>
    """, unsafe_allow_html=True)

def section_header(title, sub=None):
    st.markdown(f'<div class="section-h">{title}</div>', unsafe_allow_html=True)
    if sub:
        st.markdown(f'<div class="section-sub">{sub}</div>', unsafe_allow_html=True)


# =========================================================================
# ENGINE PERHITUNGAN (identik dgn augmented_evm_v3_boq_riskfactor.py,
# termasuk perbaikan target PlanDays per-segmen baseline)
# =========================================================================
@st.cache_data(show_spinner="Menghitung AHP, RiskFactor(t), dan simulasi Monte Carlo...")
def load_and_compute(file_evm_bytes, file_ahp_bytes, alpha, alpha_s, beta, N_SIM, seed=42):
    FILE_PATH_EVM = io.BytesIO(file_evm_bytes)
    FILE_PATH_AHP = io.BytesIO(file_ahp_bytes)

    df_umum = pd.read_excel(FILE_PATH_EVM, sheet_name="Data Umum", header=None)
    nama_proyek = str(df_umum.iloc[2, 1]) if pd.notna(df_umum.iloc[2, 1]) else "Proyek"
    BAC = float(df_umum.iloc[4, 1])
    tanggal_mulai = pd.to_datetime(df_umum.iloc[5, 1])
    PlanDaysAwal = int(df_umum.iloc[6, 1])
    PlanDaysTotal = int(df_umum.iloc[7, 1])

    df_periode = pd.read_excel(FILE_PATH_EVM, sheet_name="Data 14 Periode")
    df_periode = df_periode.dropna(subset=["Laporan Ke-"])
    T = len(df_periode)
    periode_label = ["L" + str(int(x)) for x in df_periode["Laporan Ke-"]]
    rencana_pct = df_periode["Rencana Kumulatif (%)"].to_numpy(dtype=float)
    aktual_pct = df_periode["Aktual Kumulatif (%)"].to_numpy(dtype=float)
    baseline_berlaku = df_periode["Baseline Berlaku"].to_numpy()
    tanggal_cutoff = pd.to_datetime(df_periode["Tanggal Cutoff"])
    AC = df_periode["AC Kumulatif (Rp)"].fillna(0).to_numpy(dtype=float)

    PV = rencana_pct * BAC
    EV = aktual_pct * BAC
    AD = (tanggal_cutoff - tanggal_mulai).dt.days.to_numpy().astype(float)

    mask_awal = baseline_berlaku == "Awal"
    mask_revisi = baseline_berlaku == "Revisi"
    hari_awal, rencana_awal = AD[mask_awal], rencana_pct[mask_awal]
    hari_revisi, rencana_revisi = AD[mask_revisi], rencana_pct[mask_revisi]

    def hitung_EVdays(idx, ev_pct):
        if mask_awal[idx]:
            return np.interp(ev_pct, rencana_awal, hari_awal)
        else:
            return np.interp(ev_pct, rencana_revisi, hari_revisi)

    EVdays = np.array([hitung_EVdays(t, aktual_pct[t]) for t in range(T)])
    PlanDaysTarget = np.where(mask_awal, PlanDaysAwal, PlanDaysTotal).astype(float)
    ACTUAL_FINAL_COST = AC[-1]
    ACTUAL_FINAL_DAYS = AD[-1]

    cv = EV - AC
    sv = EV - PV
    CorrFactor_value = abs(np.corrcoef(cv, sv)[0, 1])
    CorrFactor = np.full(T, CorrFactor_value)

    def ahp_weights_from_matrix(mat3x3):
        a = np.array(mat3x3, dtype=float)
        n = a.shape[0]
        K = a.prod(axis=1)
        M = K ** (1 / n)
        W = M / M.sum()
        AW = a @ W
        lam_max = np.mean(AW / W)
        CI = (lam_max - n) / (n - 1)
        RI = RI_TABLE.get(n, 1.45)
        CR = CI / RI if RI > 0 else 0.0
        return W, lam_max, CI, CR

    def _cell_to_float(v):
        if isinstance(v, str) and v.startswith("="):
            expr = v[1:].replace(" ", "")
            if "/" in expr:
                num, den = expr.split("/")
                return float(num) / float(den)
            return float(expr)
        return float(v)

    def read_ahp_level(wb, sheet_name):
        ws = wb[sheet_name]
        resp_rows = [6, 12, 18]
        mats = []
        for r0 in resp_rows:
            m = [[_cell_to_float(ws.cell(row=r0 + i, column=2 + j).value) for j in range(3)] for i in range(3)]
            mats.append(np.array(m, dtype=float))
        agg = np.ones((3, 3))
        for i in range(3):
            for j in range(3):
                agg[i, j] = np.prod([m[i, j] for m in mats]) ** (1 / 3)
        W, lam_max, CI, CR = ahp_weights_from_matrix(agg)
        return W, CR

    wb_ahp = openpyxl.load_workbook(FILE_PATH_AHP, data_only=False)
    W_kat, CR_kat = read_ahp_level(wb_ahp, "Kategori")
    W_A, CR_A = read_ahp_level(wb_ahp, "Subfaktor A")
    W_B, CR_B = read_ahp_level(wb_ahp, "Subfaktor B")
    W_C, CR_C = read_ahp_level(wb_ahp, "Subfaktor C")

    kode_label = {
        "A1": "Kesadaran efisiensi biaya perancangan", "A2": "Kompetensi personel perancangan",
        "A3": "Komunikasi & koordinasi", "B1": "Pemilihan/pengelolaan pemasok",
        "B2": "Pemilihan/pengelolaan subkontraktor", "B3": "Pengendalian biaya pengadaan",
        "C1": "Kompetensi personel manajemen lapangan", "C2": "Pengendalian mutu, harga, termin",
        "C3": "Metode pengelolaan biaya konstruksi",
    }
    kategori_label = {"A": "Perancangan (Design)", "B": "Pengadaan (Procurement)", "C": "Konstruksi (Construction)"}

    Wglobal = {
        "A1": W_kat[0] * W_A[0], "A2": W_kat[0] * W_A[1], "A3": W_kat[0] * W_A[2],
        "B1": W_kat[1] * W_B[0], "B2": W_kat[1] * W_B[1], "B3": W_kat[1] * W_B[2],
        "C1": W_kat[2] * W_C[0], "C2": W_kat[2] * W_C[1], "C3": W_kat[2] * W_C[2],
    }
    ahp_summary = {
        "Kategori": {"W": W_kat, "CR": CR_kat, "labels": ["A - Perancangan", "B - Pengadaan", "C - Konstruksi"]},
        "Subfaktor A": {"W": W_A, "CR": CR_A, "labels": ["A1", "A2", "A3"]},
        "Subfaktor B": {"W": W_B, "CR": CR_B, "labels": ["B1", "B2", "B3"]},
        "Subfaktor C": {"W": W_C, "CR": CR_C, "labels": ["C1", "C2", "C3"]},
    }
    Wkat_dict = {"A": W_kat[0], "B": W_kat[1], "C": W_kat[2]}

    def read_boq_sections(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["BOQ"]
        section_rows = {}
        roman = ["I", "II", "III", "IV", "V"]
        for r in range(1, ws.max_row + 1):
            b = ws.cell(row=r, column=2).value
            if isinstance(b, str) and b.strip() in roman:
                section_rows[b.strip()] = r
        sections = {}
        used_fallback = False
        for k in roman:
            jumlah_row = None
            for r in range(section_rows[k] + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=2).value
                if isinstance(b, str) and b.strip() == f"JUMLAH {k}":
                    jumlah_row = r
                    break
            material = ws.cell(row=jumlah_row, column=18).value
            jasa = ws.cell(row=jumlah_row, column=19).value
            if material is None and jasa is None:
                material = _BOQ_FALLBACK[k]["material"]
                jasa = _BOQ_FALLBACK[k]["jasa"]
                used_fallback = True
            else:
                material = material or 0
                jasa = jasa or 0
            sections[k] = {"material": float(material), "jasa": float(jasa)}
            sections[k]["total"] = sections[k]["material"] + sections[k]["jasa"]
        return sections, used_fallback

    FILE_PATH_EVM.seek(0)
    boq_sections, used_fallback = read_boq_sections(FILE_PATH_EVM)
    grand_total = sum(s["total"] for s in boq_sections.values())

    cum = 0.0
    bounds = {}
    for k in ["I", "II", "III", "IV", "V"]:
        start = cum
        cum += boq_sections[k]["total"] / grand_total
        bounds[k] = (start, cum)

    matII = boq_sections["II"]["material"]
    jasII = boq_sections["II"]["jasa"]
    fB_II = matII / (matII + jasII)
    fC_II = jasII / (matII + jasII)
    C1w, C2w, C3w = W_C
    renorm13 = (C1w / (C1w + C3w), C3w / (C1w + C3w))

    def overlap(a0, a1, b0, b1):
        return max(0.0, min(a1, b1) - max(a0, b0))

    sections_order = ["I", "II", "III", "IV", "V"]
    section_share = np.zeros((T, 5))
    for t in range(T):
        p0 = aktual_pct[t - 1] if t > 0 else 0.0
        p1 = aktual_pct[t]
        dP = p1 - p0
        if dP <= 1e-12:
            continue
        for si, s in enumerate(sections_order):
            b0, b1 = bounds[s]
            section_share[t, si] = overlap(p0, p1, b0, b1) / dP

    subfactor_share = {k: np.zeros(T) for k in Wglobal}
    for t in range(T):
        sI, sII, sIII, sIV, sV = section_share[t]
        for k in ["A1", "A2", "A3"]:
            subfactor_share[k][t] += sI / 3
        for k in ["B1", "B2", "B3"]:
            subfactor_share[k][t] += sII * fB_II / 3
        for k in ["C1", "C3"]:
            subfactor_share[k][t] += (sII * fC_II + sIII + sV) * (renorm13[0] if k == "C1" else renorm13[1])
        subfactor_share["C2"][t] += sIV

    RiskFactor_BOQ = np.full(T, np.nan)
    for t in range(T):
        tot = sum(subfactor_share[k][t] for k in subfactor_share)
        if tot > 1e-12:
            RiskFactor_BOQ[t] = sum((subfactor_share[k][t] / tot) * Wglobal[k] for k in subfactor_share)
    last_valid = 0.05
    for t in range(T):
        if np.isnan(RiskFactor_BOQ[t]):
            RiskFactor_BOQ[t] = last_valid
        else:
            last_valid = RiskFactor_BOQ[t]

    def _cpi_spi_traditional():
        cpi = np.where((AC > 0) & (EV > 0), EV / AC, 1.0)
        spi = np.where(AD != 0, EVdays / AD, 1.0)
        costtrad = np.array([AC[t] + (BAC - EV[t]) / cpi[t] if cpi[t] != 0 else np.nan for t in range(T)])
        daystrad = np.array([AD[t] + (PlanDaysTarget[t] - EVdays[t]) / spi[t] if spi[t] != 0 else np.nan for t in range(T)])
        return costtrad, daystrad

    def _pct_change(series):
        out = np.full(T, np.nan)
        idxs = np.where(np.isfinite(series))[0]
        for k in range(1, len(idxs)):
            i0, i1 = idxs[k - 1], idxs[k]
            out[i1] = abs(series[i1] - series[i0]) / abs(series[i0])
        return out

    CostTrad, DaysTrad = _cpi_spi_traditional()
    _vol_cost = _pct_change(CostTrad)
    _vol_days = _pct_change(DaysTrad)
    _vol = np.nanmean(np.vstack([_vol_cost, _vol_days]), axis=0)
    _vol = np.where(np.isnan(_vol), np.nanmin(_vol), _vol)
    VolatilityFactor = np.clip(_vol / 0.30, 0, 1)

    RiskFactor_boq_norm = RiskFactor_BOQ / RiskFactor_BOQ.max()
    RiskFactor = 0.5 * RiskFactor_boq_norm + 0.5 * VolatilityFactor
    RiskFactor = RiskFactor * RiskFactor_BOQ.max() / RiskFactor_boq_norm[RiskFactor_BOQ.argmax()]

    def AugmentedEVM(N, seed=42):
        rng = np.random.default_rng(seed)
        Results = {}
        for t in range(T):
            CPI_t = EV[t] / AC[t] if AC[t] > 0 and EV[t] > 0 else 1.0
            CostTrad_t = AC[t] + (BAC - EV[t]) / CPI_t
            SPI_t = EVdays[t] / AD[t] if AD[t] != 0 else 1.0
            DaysTrad_t = AD[t] + (PlanDaysTarget[t] - EVdays[t]) / SPI_t if SPI_t != 0 else PlanDaysTarget[t]
            sigma_t = beta * RiskFactor[t]
            Z_n = rng.normal(1.0, sigma_t, N)
            costDist = CostTrad_t * (1 + CorrFactor[t] * alpha) * Z_n
            Y_n = rng.normal(1.0, sigma_t, N)
            daysDist = DaysTrad_t * (1 + CorrFactor[t] * alpha_s) * Y_n
            Results[t] = {
                "CPI": CPI_t, "SPI": SPI_t, "CostTrad": CostTrad_t, "DaysTrad": DaysTrad_t,
                "CostAugMean": costDist.mean(), "CostAugP5": np.percentile(costDist, 5), "CostAugP95": np.percentile(costDist, 95),
                "DaysAugMean": daysDist.mean(), "DaysAugP5": np.percentile(daysDist, 5), "DaysAugP95": np.percentile(daysDist, 95),
                "CostDistArray": costDist, "DaysDistArray": daysDist,
            }
        return Results

    Results = AugmentedEVM(N_SIM, seed)

    CostTradArr = np.array([Results[t]["CostTrad"] for t in range(T)])
    DaysTradArr = np.array([Results[t]["DaysTrad"] for t in range(T)])
    CostAugMean = np.array([Results[t]["CostAugMean"] for t in range(T)])
    CostAugP5 = np.array([Results[t]["CostAugP5"] for t in range(T)])
    CostAugP95 = np.array([Results[t]["CostAugP95"] for t in range(T)])
    DaysAugMean = np.array([Results[t]["DaysAugMean"] for t in range(T)])
    DaysAugP5 = np.array([Results[t]["DaysAugP5"] for t in range(T)])
    DaysAugP95 = np.array([Results[t]["DaysAugP95"] for t in range(T)])

    df_hasil = pd.DataFrame({
        "Periode": periode_label, "RiskFactor": RiskFactor, "RiskFactor_BOQ": RiskFactor_BOQ,
        "Volatilitas": VolatilityFactor, "CorrFactor": CorrFactor,
        "PV": PV, "EV": EV, "AC": AC, "AD": AD, "EVdays": EVdays,
        "CostTrad": CostTradArr, "CostAugMean": CostAugMean, "CostAugP5": CostAugP5, "CostAugP95": CostAugP95,
        "DaysTrad": DaysTradArr, "DaysAugMean": DaysAugMean, "DaysAugP5": DaysAugP5, "DaysAugP95": DaysAugP95,
        "Baseline": baseline_berlaku,
    })

    valid_cost = np.isfinite(CostTradArr)
    cov_cost = (CostAugP5 <= ACTUAL_FINAL_COST) & (ACTUAL_FINAL_COST <= CostAugP95) & valid_cost
    n_valid_cost = int(valid_cost.sum())
    is_revisi = baseline_berlaku == "Revisi"
    valid_days_all = np.isfinite(DaysTradArr)
    cov_days_all = (DaysAugP5 <= ACTUAL_FINAL_DAYS) & (ACTUAL_FINAL_DAYS <= DaysAugP95) & valid_days_all
    cov_days_revisi = cov_days_all & is_revisi
    n_valid_days_revisi = int((valid_days_all & is_revisi).sum())

    mape_cost_mask = valid_cost & (np.arange(T) >= 2)
    mape_cost = np.mean(np.abs(ACTUAL_FINAL_COST - CostTradArr[mape_cost_mask]) / ACTUAL_FINAL_COST) * 100
    mape_days_mask = valid_days_all & (np.arange(T) >= 2)
    mape_days = np.mean(np.abs(ACTUAL_FINAL_DAYS - DaysTradArr[mape_days_mask]) / ACTUAL_FINAL_DAYS) * 100

    validasi = {
        "cov_cost_pct": cov_cost.sum() / n_valid_cost * 100 if n_valid_cost else np.nan,
        "cov_cost_n": (int(cov_cost.sum()), n_valid_cost),
        "cov_days_pct": cov_days_revisi.sum() / n_valid_days_revisi * 100 if n_valid_days_revisi else np.nan,
        "cov_days_n": (int(cov_days_revisi.sum()), n_valid_days_revisi),
        "mape_cost": mape_cost, "mape_days": mape_days,
    }

    return {
        "nama_proyek": nama_proyek, "T": T, "periode_label": periode_label, "BAC": BAC,
        "PlanDaysAwal": PlanDaysAwal, "PlanDaysTotal": PlanDaysTotal,
        "rencana_pct": rencana_pct, "aktual_pct": aktual_pct, "PV": PV, "EV": EV, "AC": AC, "AD": AD,
        "EVdays": EVdays, "ACTUAL_FINAL_COST": ACTUAL_FINAL_COST, "ACTUAL_FINAL_DAYS": ACTUAL_FINAL_DAYS,
        "CorrFactor_value": CorrFactor_value, "Wglobal": Wglobal, "ahp_summary": ahp_summary,
        "kode_label": kode_label, "kategori_label": kategori_label, "Wkat_dict": Wkat_dict,
        "section_share": section_share, "sections_order": sections_order, "boq_sections": boq_sections,
        "RiskFactor": RiskFactor, "RiskFactor_BOQ": RiskFactor_BOQ, "VolatilityFactor": VolatilityFactor,
        "Results": Results, "df_hasil": df_hasil, "validasi": validasi, "used_boq_fallback": used_fallback,
        "baseline_berlaku": baseline_berlaku,
    }


def rp(x, unit=True):
    if abs(x) >= 1e9:
        return f"Rp{x/1e9:,.2f} M" if unit else f"{x/1e9:,.2f}"
    return f"Rp{x/1e6:,.0f} jt" if unit else f"{x/1e6:,.0f}"


# =========================================================================
# SIDEBAR
# =========================================================================
with st.sidebar:
    if LOGO_B64:
        st.markdown(f"""
        <div class="sidebar-institution">
          <img src="data:image/png;base64,{LOGO_B64}" alt="Logo {UNIVERSITAS}" />
          <div>
            <div class="inst-name">{UNIVERSITAS}</div>
            <div class="inst-faculty">{FAKULTAS}</div>
            <div class="inst-jurusan">{JURUSAN}</div>
          </div>
        </div>
        """, unsafe_allow_html=True)
    if PRIME_LOGO_B64:
        st.markdown(f"""
        <div class="sidebar-prime">
          <img src="data:image/png;base64,{PRIME_LOGO_B64}" alt="Logo {PRIME_NAME}" />
          <div>
            <div class="prime-word">{PRIME_NAME}</div>
            <div class="prime-tag">{PRIME_TAGLINE}</div>
          </div>
        </div>
        """, unsafe_allow_html=True)
    st.markdown('<div class="sidebar-brand">RISK-BASED EVM ANALYZER</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="sidebar-caption">Sistem evaluasi kinerja biaya dan waktu proyek EPC '
        'berbasis risiko, mengintegrasikan AHP, Earned Value Management, dan simulasi Monte Carlo.</div>',
        unsafe_allow_html=True,
    )
    st.markdown("---")
    st.markdown('<div class="sidebar-step">Langkah 1 — Unggah Data Proyek</div>', unsafe_allow_html=True)
    st.caption(
        "Belum punya file dengan format yang sesuai? Unduh template kosong di bawah, "
        "isi dengan data proyek Anda sendiri, lalu unggah kembali di sini."
    )
    tcol1, tcol2 = st.columns(2)
    with tcol1:
        st.download_button(
            "Template EVM", build_evm_template_bytes(), "Template_Data_EVM.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, help="Sheet Data Umum, Data 14 Periode, BOQ - siap isi",
        )
    with tcol2:
        st.download_button(
            "Template AHP", build_ahp_template_bytes(), "Template_Perhitungan_AHP.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, help="Matriks perbandingan berpasangan 3 responden - siap isi",
        )
    file_evm = st.file_uploader("File data EVM (.xlsx)", type=["xlsx"], help="Berisi sheet Data Umum, Data 14 Periode, BOQ")
    file_ahp = st.file_uploader("File perhitungan AHP (.xlsx)", type=["xlsx"], help="Berisi matriks perbandingan berpasangan 3 responden")

    st.markdown('<div class="sidebar-step">Langkah 2 — Parameter Model</div>', unsafe_allow_html=True)
    N_SIM = st.select_slider("Jumlah iterasi Monte Carlo (N)", options=[100, 500, 1000, 2500, 5000, 10000, 25000, 50000], value=10000)
    beta = st.slider("β — koefisien penyebaran risiko", 0.1, 2.0, 1.0, 0.1)
    alpha = st.slider("α — pengaruh CorrFactor (biaya)", 0.1, 1.0, 0.5, 0.1)
    alpha_s = st.slider("α_s — pengaruh CorrFactor (waktu)", 0.1, 1.0, 0.5, 0.1)

    st.markdown('<div class="sidebar-step">Langkah 3 — Jalankan Analisis</div>', unsafe_allow_html=True)
    st.caption(
        "Perhitungan **tidak** berjalan otomatis saat Anda menggeser slider di atas — "
        "atur dulu semua parameter, baru klik tombol ini untuk memproses."
    )
    run_clicked = st.button("Jalankan Program", type="primary", use_container_width=True)

    st.markdown("---")
    with st.expander("Tentang Sistem Ini"):
        st.markdown(
            "Dashboard ini menerapkan kerangka **Risk-Based Earned Value Management** yang dapat "
            "dipakai pada proyek EPC mana pun, selama tersedia: (1) data historis EVM per periode "
            "(PV/EV/AC), (2) Bill of Quantity, dan (3) hasil kuesioner AHP dari pakar proyek. "
            "Metodologi: AHP → bobot risiko → RiskFactor(t) dinamis → simulasi Monte Carlo → "
            "Augmented EVM → validasi model."
        )

if file_evm is None or file_ahp is None:
    # =====================================================================
    # LANDING PAGE (belum ada data)
    # =====================================================================
    st.markdown(_brand_ribbon_html(), unsafe_allow_html=True)
    st.markdown(f"""
    <div class="topbar"><span><span class="dot">●</span>Sistem Analisis Kinerja Proyek</span><span>Risk-Based EVM</span></div>
    <div class="page-head">
      <div class="eyebrow">Risk-Based Earned Value Management</div>
      <h1>Analisis Kinerja Proyek EPC yang Memperhitungkan Risiko Nyata di Lapangan</h1>
      <p>Dashboard ini mengintegrasikan pembobotan risiko (AHP), evaluasi kinerja biaya dan waktu
      (Earned Value Management), dan simulasi probabilistik (Monte Carlo) menjadi satu alat
      analisis yang dapat diterapkan pada proyek Engineering, Procurement, and Construction (EPC)
      secara umum, tidak terbatas pada proyek yang datanya sedang dimuat saat ini.</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-h">Alur Kerja Sistem</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    steps = [
        ("1", "Unggah Data", "Masukkan data historis EVM (PV/EV/AC per periode, BOQ) dan hasil kuesioner AHP dari pakar proyek Anda."),
        ("2", "Bobot Risiko (AHP)", "Sistem menghitung bobot prioritas tiap faktor risiko dan menguji konsistensi penilaian (CR)."),
        ("3", "RiskFactor(t) Dinamis", "Bobot risiko diaktivasi mengikuti pekerjaan yang sedang berjalan tiap periode, dipadukan sinyal volatilitas historis."),
        ("4", "Simulasi & Validasi", "Monte Carlo menghasilkan sebaran probabilistik biaya & waktu akhir, divalidasi terhadap capaian aktual."),
    ]
    for col, (num, title, desc) in zip([c1, c2, c3, c4], steps):
        with col:
            st.markdown(f"""
            <div class="step-card">
              <div class="step-num">{num}</div>
              <h4>{title}</h4>
              <p>{desc}</p>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    insight_box(
        "Cocok untuk proyek seperti apa?",
        "Kerangka ini berlaku umum untuk proyek EPC berskala menengah-besar yang memiliki laporan "
        "progres periodik, rincian Bill of Quantity, dan akses ke tenaga ahli (Project Manager, "
        "Site Engineer, Project Control) untuk pengisian kuesioner AHP. Contoh data yang menyertai "
        "alat ini berasal dari studi kasus proyek perbaikan tangki timbun berskema kontrak EPC."
    )
    st.info("Unggah kedua file Excel pada panel kiri untuk mulai menganalisis proyek Anda.")
    st.stop()

# =========================================================================
# MUAT & HITUNG DATA — HANYA saat tombol "Jalankan Program" diklik.
# Hasil disimpan di session_state supaya interaksi lain (buka tab, dst.)
# tidak memicu ulang simulasi Monte Carlo yang berat.
# =========================================================================
current_params = (file_evm.name, file_evm.size, file_ahp.name, file_ahp.size, alpha, alpha_s, beta, N_SIM)

if run_clicked:
    with st.spinner("Menghitung AHP, RiskFactor(t), dan simulasi Monte Carlo..."):
        st.session_state["evm_data"] = load_and_compute(
            file_evm.getvalue(), file_ahp.getvalue(), alpha, alpha_s, beta, N_SIM
        )
        st.session_state["evm_params"] = current_params

if "evm_data" not in st.session_state:
    st.markdown(_brand_ribbon_html(), unsafe_allow_html=True)
    st.info("Data sudah diunggah. Atur parameter model di sidebar (opsional), lalu klik **Jalankan Program** untuk memulai analisis.")
    st.stop()

if st.session_state.get("evm_params") != current_params:
    st.warning(
        "Parameter atau file telah diubah sejak analisis terakhir dijalankan. "
        "Hasil di bawah masih menampilkan run sebelumnya — klik **Jalankan Program** lagi untuk memperbarui."
    )

data = st.session_state["evm_data"]
T = data["T"]; labels = data["periode_label"]; x = list(range(T))
df = data["df_hasil"]; v = data["validasi"]

if data["used_boq_fallback"]:
    st.warning(
        "Cache formula BOQ pada file Excel kosong (belum di-refresh sejak terakhir disimpan). "
        "Dashboard memakai nilai cadangan terverifikasi - buka & simpan ulang file di Microsoft "
        "Excel untuk menyegarkan cache."
    )

# =========================================================================
# HERO PROYEK AKTIF
# =========================================================================
cost_variance_pct = (data["ACTUAL_FINAL_COST"] - data["BAC"]) / data["BAC"] * 100
schedule_variance_days = data["ACTUAL_FINAL_DAYS"] - data["PlanDaysTotal"]
cost_status = ("Di Bawah Anggaran", "badge-good") if cost_variance_pct <= 0 else ("Melebihi Anggaran", "badge-bad")
sched_status = ("Tepat / Lebih Cepat", "badge-good") if schedule_variance_days <= 0 else ("Terlambat", "badge-bad")

st.markdown(_brand_ribbon_html(), unsafe_allow_html=True)
st.markdown(f"""
<div class="topbar"><span><span class="dot">●</span>Sistem Analisis Kinerja Proyek</span><span>Risk-Based EVM</span></div>
<div class="page-head">
  <div class="eyebrow">Proyek Aktif Dianalisis</div>
  <h1>{data['nama_proyek']}</h1>
  <p>
    <span class="badge {cost_status[1]}">BIAYA: {cost_status[0]} ({cost_variance_pct:+.1f}%)</span>
    &nbsp;&nbsp;
    <span class="badge {sched_status[1]}">JADWAL: {sched_status[0]} ({schedule_variance_days:+.0f} hari)</span>
  </p>
</div>
""", unsafe_allow_html=True)

tabs = st.tabs([
    "1. Ringkasan Eksekutif", "2. Profil Proyek", "3. Analisis Risiko (AHP)",
    "4. Kinerja EVM", "5. Model Augmented EVM", "6. Validasi Model",
])

# =========================================================================
# TAB 0: RINGKASAN EKSEKUTIF
# =========================================================================
with tabs[0]:
    section_header("Angka Kunci")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        metric_card("Anggaran (BAC)", rp(data["BAC"]))
    with c2:
        cost_pct_used = data["ACTUAL_FINAL_COST"] / data["BAC"] * 100
        metric_card("Biaya Aktual Akhir", rp(data["ACTUAL_FINAL_COST"]),
                     f"{cost_variance_pct:+.1f}% dari BAC", C["bad"] if cost_variance_pct > 0 else C["good"],
                     pct=cost_pct_used, pct_color=C["bad"] if cost_variance_pct > 0 else C["good"])
    with c3:
        metric_card("Durasi Rencana", f"{data['PlanDaysTotal']:.0f} hari")
    with c4:
        days_pct_used = data["ACTUAL_FINAL_DAYS"] / data["PlanDaysTotal"] * 100
        metric_card("Durasi Aktual", f"{data['ACTUAL_FINAL_DAYS']:.0f} hari",
                     f"{schedule_variance_days:+.0f} hari dari rencana", C["bad"] if schedule_variance_days > 0 else C["good"],
                     pct=days_pct_used, pct_color=C["bad"] if schedule_variance_days > 0 else C["good"])

    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    top_subfaktor = max(data["Wglobal"], key=data["Wglobal"].get)
    top_kat = max(data["Wkat_dict"], key=data["Wkat_dict"].get)
    with c1:
        metric_card("Faktor Risiko Dominan", top_subfaktor, data["kode_label"][top_subfaktor][:28] + "...")
    with c2:
        metric_card("Kategori Risiko Utama", data["kategori_label"][top_kat].split(" (")[0],
                     f"{data['Wkat_dict'][top_kat]*100:.1f}% dari total bobot risiko")
    with c3:
        metric_card("Cakupan Prediksi Biaya", f"{v['cov_cost_pct']:.0f}%",
                     f"{v['cov_cost_n'][0]}/{v['cov_cost_n'][1]} periode tercakup")
    with c4:
        metric_card("Cakupan Prediksi Waktu", f"{v['cov_days_pct']:.0f}%",
                     f"{v['cov_days_n'][0]}/{v['cov_days_n'][1]} periode tercakup")

    st.markdown("<br>", unsafe_allow_html=True)
    section_header("Apa Makna Hasil Ini?", "Interpretasi otomatis berdasarkan angka yang dihitung dari data Anda")

    # --- Insight 1: status akhir proyek ---
    if cost_variance_pct <= 0 and schedule_variance_days <= 0:
        ringkas = "proyek ini berhasil diselesaikan <b>sesuai atau di bawah anggaran, dan tepat/lebih cepat dari jadwal</b>"
    elif cost_variance_pct <= 0 and schedule_variance_days > 0:
        ringkas = f"proyek ini terkendali dari sisi biaya, namun <b>mengalami keterlambatan {schedule_variance_days:.0f} hari</b> dari target"
    elif cost_variance_pct > 0 and schedule_variance_days <= 0:
        ringkas = f"proyek ini selesai tepat/lebih cepat dari jadwal, namun <b>melebihi anggaran sebesar {cost_variance_pct:.1f}%</b>"
    else:
        ringkas = f"proyek ini <b>mengalami pembengkakan biaya {cost_variance_pct:.1f}%</b> sekaligus <b>keterlambatan {schedule_variance_days:.0f} hari</b> dari target"

    insight_box(
        "Status Akhir Proyek",
        f"Berdasarkan perbandingan realisasi terhadap rencana, {ringkas}. "
        f"Anggaran (BAC) sebesar {rp(data['BAC'])} terealisasi menjadi {rp(data['ACTUAL_FINAL_COST'])}, "
        f"dan durasi rencana {data['PlanDaysTotal']:.0f} hari terealisasi menjadi {data['ACTUAL_FINAL_DAYS']:.0f} hari."
    )

    # --- Insight 2: AHP ---
    top3 = sorted(data["Wglobal"].items(), key=lambda kv: -kv[1])[:3]
    top3_txt = ", ".join([f"<b>{k}</b> ({data['kode_label'][k]}, {w*100:.1f}%)" for k, w in top3])
    insight_box(
        "Faktor Risiko yang Paling Menentukan",
        f"Hasil pembobotan AHP dari penilaian para ahli menunjukkan bahwa risiko proyek ini paling "
        f"didominasi oleh kategori <b>{data['kategori_label'][top_kat]}</b> ({data['Wkat_dict'][top_kat]*100:.1f}% "
        f"dari seluruh bobot risiko). Tiga subfaktor risiko berkontribusi terbesar: {top3_txt}. "
        f"Ini mengindikasikan bahwa pengendalian risiko proyek sebaiknya diprioritaskan pada aspek-aspek tersebut."
    )

    # --- Insight 3: tren risiko waktu ---
    peak_t = int(np.argmax(data["RiskFactor"]))
    trend_awal = data["RiskFactor"][0]
    trend_akhir = data["RiskFactor"][-1]
    arah = "meningkat" if trend_akhir > trend_awal else "menurun"
    insight_box(
        "Bagaimana Risiko Berubah Sepanjang Waktu",
        f"Tingkat risiko proyek (RiskFactor) secara umum <b>{arah}</b> dari periode awal ({trend_awal:.3f}) "
        f"ke periode akhir ({trend_akhir:.3f}), dengan puncak risiko tertinggi terjadi pada periode "
        f"<b>{labels[peak_t]}</b> ({data['RiskFactor'][peak_t]:.3f}). Periode ini layak menjadi perhatian "
        f"khusus tim pengendali proyek karena kombinasi jenis pekerjaan yang sedang berjalan dan "
        f"volatilitas kinerja biaya-waktu historis berada pada titik tertinggi."
    )

    # --- Insight 4: keandalan model ---
    kualitas_cov_cost = "baik" if v["cov_cost_pct"] >= 70 else "perlu perhatian"
    insight_box(
        "Seberapa Bisa Diandalkan Model Ini?",
        f"Model Augmented EVM berhasil mencakup nilai aktual akhir proyek dalam rentang prediksinya "
        f"(P5-P95) pada <b>{v['cov_cost_pct']:.0f}% periode</b> untuk biaya dan <b>{v['cov_days_pct']:.0f}% "
        f"periode</b> untuk waktu (dibandingkan pada periode dengan target rencana yang sama) - "
        f"tingkat keandalan yang <b>{kualitas_cov_cost}</b>. Sementara itu, estimasi EVM tradisional "
        f"(tanpa penyesuaian risiko) memiliki rata-rata kesalahan (MAPE) sebesar "
        f"<b>{v['mape_cost']:.1f}%</b> untuk biaya dan <b>{v['mape_days']:.1f}%</b> untuk waktu - "
        f"menunjukkan bahwa pendekatan berbasis risiko memberikan gambaran ketidakpastian yang "
        f"lebih realistis dibandingkan estimasi deterministik konvensional."
    )

    section_header("Kurva S Proyek", "Progres rencana vs realisasi")
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=x, y=data["rencana_pct"] * 100, name="Rencana Kumulatif (%)",
                              line=dict(color=C["navy"], width=3)))
    fig.add_trace(go.Scatter(x=x, y=data["aktual_pct"] * 100, name="Aktual Kumulatif (%)",
                              line=dict(color=C["amber"], width=3)))
    fig.update_xaxes(tickmode="array", tickvals=x, ticktext=labels)
    fig.update_yaxes(title="Progres Kumulatif (%)")
    st.plotly_chart(style_fig(fig, height=380), use_container_width=True)

# =========================================================================
# TAB 1: PROFIL PROYEK
# =========================================================================
with tabs[1]:
    section_header("Data Umum Proyek")
    c1, c2 = st.columns(2)
    with c1:
        metric_card("Anggaran (BAC)", rp(data["BAC"]))
        st.markdown("<br>", unsafe_allow_html=True)
        metric_card("Durasi Rencana Awal", f"{data['PlanDaysAwal']:.0f} hari")
    with c2:
        metric_card("Jumlah Periode Evaluasi", f"{T} periode")
        st.markdown("<br>", unsafe_allow_html=True)
        metric_card("Durasi Rencana Revisi", f"{data['PlanDaysTotal']:.0f} hari")

    section_header("Biaya & Jadwal: Rencana vs Aktual")
    c1, c2 = st.columns(2)
    with c1:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=x, y=data["PV"], name="Planned Value (PV)", line=dict(color=C["navy"], width=2.5)))
        fig.add_trace(go.Scatter(x=x, y=data["AC"], name="Actual Cost (AC)", line=dict(color=C["amber"], width=2.5)))
        fig.update_xaxes(tickmode="array", tickvals=x, ticktext=labels)
        fig.update_yaxes(title="Rupiah")
        st.plotly_chart(style_fig(fig, height=360, title="Biaya"), use_container_width=True)
    with c2:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=x, y=data["AD"], name="Planned Days (AD)", line=dict(color=C["navy"], width=2.5)))
        fig.add_trace(go.Scatter(x=x, y=data["EVdays"], name="Actual Progress (EVdays)", line=dict(color=C["amber"], width=2.5)))
        fig.update_xaxes(tickmode="array", tickvals=x, ticktext=labels)
        fig.update_yaxes(title="Hari")
        st.plotly_chart(style_fig(fig, height=360, title="Jadwal"), use_container_width=True)

    section_header("Tabel Data per Periode")
    st.dataframe(
        df[["Periode", "Baseline", "PV", "EV", "AC", "AD", "EVdays"]]
        .style.format({"PV": "{:,.0f}", "EV": "{:,.0f}", "AC": "{:,.0f}", "AD": "{:.0f}", "EVdays": "{:.1f}"}),
        use_container_width=True, hide_index=True,
    )

# =========================================================================
# TAB 2: ANALISIS RISIKO (AHP)
# =========================================================================
with tabs[2]:
    section_header("Uji Konsistensi Penilaian", "Consistency Ratio (CR) harus < 0,10 agar bobot dapat dipakai")
    cols = st.columns(4)
    for i, (name, d) in enumerate(data["ahp_summary"].items()):
        with cols[i]:
            ok = d["CR"] < 0.1
            badge = "badge-good" if ok else "badge-bad"
            st.markdown(f"""
            <div class="metric-card">
              <div class="metric-label">{name}</div>
              <div class="metric-value">{d['CR']:.4f}</div>
              <div class="metric-sub"><span class="badge {badge}">{'KONSISTEN' if ok else 'PERLU REVISI'}</span></div>
            </div>
            """, unsafe_allow_html=True)

    all_ok = all(d["CR"] < 0.1 for d in data["ahp_summary"].values())
    insight_box(
        "Interpretasi Uji Konsistensi",
        ("Seluruh level penilaian <b>konsisten</b> (CR &lt; 0,10) - bobot prioritas risiko yang "
         "dihasilkan layak digunakan sebagai dasar pengambilan keputusan." if all_ok else
         "Terdapat level penilaian yang <b>belum konsisten</b> (CR ≥ 0,10) - disarankan meninjau "
         "ulang matriks perbandingan berpasangan pada level tersebut bersama responden terkait "
         "sebelum bobot dipakai lebih lanjut.")
    )

    section_header("Bobot Global 9 Subfaktor Risiko")
    items = sorted(data["Wglobal"].items(), key=lambda kv: kv[1])
    codes = [k for k, v_ in items]
    vals = [v_ for k, v_ in items]
    colors = [C["amber"] if k == top_subfaktor else C["navy2"] for k in codes]
    fig = go.Figure(go.Bar(
        x=vals, y=[f"{k} — {data['kode_label'][k]}" for k in codes], orientation="h",
        marker_color=colors, text=[f"{v_*100:.1f}%" for v_ in vals], textposition="outside",
    ))
    fig.update_xaxes(title="Bobot Global", range=[0, max(vals) * 1.25])
    st.plotly_chart(style_fig(fig, height=430), use_container_width=True)

    section_header("Bobot per Level Hierarki")
    for name, d in data["ahp_summary"].items():
        st.markdown(f"**{name}**")
        dfw = pd.DataFrame({"Subfaktor": d["labels"], "Bobot": d["W"]})
        st.dataframe(dfw.style.format({"Bobot": "{:.4f}"}), use_container_width=True, hide_index=True)

# =========================================================================
# TAB 3: KINERJA EVM
# =========================================================================
with tabs[3]:
    section_header("Indeks Kinerja per Periode")
    df_show = df.copy()
    df_show["CPI"] = df_show["EV"] / df_show["AC"].replace(0, np.nan)
    df_show["SPI"] = df_show["EV"] / df_show["PV"].replace(0, np.nan)
    st.dataframe(
        df_show[["Periode", "Baseline", "CPI", "SPI", "CostTrad", "DaysTrad"]]
        .style.format({"CPI": "{:.3f}", "SPI": "{:.3f}", "CostTrad": "{:,.0f}", "DaysTrad": "{:.1f}"}),
        use_container_width=True, hide_index=True,
    )

    avg_cpi = df_show["CPI"].mean()
    avg_spi = df_show["SPI"].mean()
    cpi_ket = "efisien (CPI>1)" if avg_cpi > 1 else "kurang efisien (CPI<1)"
    spi_ket = "lebih cepat dari rencana (SPI>1)" if avg_spi > 1 else "lebih lambat dari rencana (SPI<1)"
    insight_box(
        "Interpretasi Kinerja Tradisional",
        f"Rata-rata CPI sepanjang periode evaluasi adalah <b>{avg_cpi:.2f}</b> ({cpi_ket}), dan rata-rata "
        f"SPI adalah <b>{avg_spi:.2f}</b> ({spi_ket}). Nilai <b>CorrFactor = {data['CorrFactor_value']:.4f}</b> "
        f"menunjukkan tingkat keterkaitan historis antara penyimpangan biaya dan penyimpangan jadwal proyek ini."
    )

# =========================================================================
# TAB 4: MODEL AUGMENTED EVM
# =========================================================================
with tabs[4]:
    section_header("RiskFactor(t) - Dinamika Risiko per Periode")
    c1, c2 = st.columns([3, 2])
    with c1:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=x, y=data["RiskFactor_BOQ"], name="BOQ Aktif murni",
                                  line=dict(color=C["navy"], width=2, dash="dash")))
        fig.add_trace(go.Scatter(x=x, y=data["RiskFactor"], name="Gabungan (BOQ + Volatilitas)",
                                  line=dict(color=C["red"], width=3)))
        fig.update_xaxes(tickmode="array", tickvals=x, ticktext=labels)
        st.plotly_chart(style_fig(fig, height=380), use_container_width=True)
    with c2:
        section_names = ["I. Persiapan", "II. Perbaikan Tangki", "III. Pengecatan", "IV. Pengetesan", "V. Finishing"]
        fig = go.Figure(go.Heatmap(
            z=data["section_share"].T, x=labels, y=section_names,
            colorscale=[[0, "#FDF6EE"], [1, C["amber"]]], showscale=True,
        ))
        st.plotly_chart(style_fig(fig, height=380), use_container_width=True)

    section_header("Forecast Biaya & Waktu", "Garis putus-putus = EVM tradisional, garis merah = rata-rata Augmented EVM, area = selang P5-P95")
    c1, c2 = st.columns(2)
    with c1:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=x + x[::-1], y=list(df["CostAugP95"]) + list(df["CostAugP5"])[::-1],
                                  fill="toself", fillcolor="rgba(30,111,166,0.15)", line=dict(width=0), name="P5-P95", showlegend=True))
        fig.add_hline(y=data["ACTUAL_FINAL_COST"], line=dict(color="black", dash="dot"), annotation_text="Actual Final")
        fig.add_trace(go.Scatter(x=x, y=df["CostTrad"], name="EVM Tradisional", line=dict(color=C["navy"], width=2, dash="dash")))
        fig.add_trace(go.Scatter(x=x, y=df["CostAugMean"], name="Augmented (mean)", line=dict(color=C["red"], width=3)))
        fig.update_xaxes(tickmode="array", tickvals=x, ticktext=labels)
        st.plotly_chart(style_fig(fig, height=380, title="Biaya"), use_container_width=True)
    with c2:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=x + x[::-1], y=list(df["DaysAugP95"]) + list(df["DaysAugP5"])[::-1],
                                  fill="toself", fillcolor="rgba(30,111,166,0.15)", line=dict(width=0), name="P5-P95", showlegend=True))
        fig.add_hline(y=data["ACTUAL_FINAL_DAYS"], line=dict(color="black", dash="dot"), annotation_text="Actual Final")
        fig.add_trace(go.Scatter(x=x, y=df["DaysTrad"], name="EVM Tradisional", line=dict(color=C["navy"], width=2, dash="dash")))
        fig.add_trace(go.Scatter(x=x, y=df["DaysAugMean"], name="Augmented (mean)", line=dict(color=C["red"], width=3)))
        fig.update_xaxes(tickmode="array", tickvals=x, ticktext=labels)
        st.plotly_chart(style_fig(fig, height=380, title="Waktu"), use_container_width=True)

    section_header("Distribusi Monte Carlo per Periode")
    c1, c2 = st.columns(2)
    with c1:
        fig = go.Figure()
        for t in range(T):
            fig.add_trace(go.Violin(y=data["Results"][t]["CostDistArray"], x0=labels[t], name=labels[t],
                                     line_color=C["accent"], fillcolor="rgba(30,111,166,0.25)", showlegend=False, points=False))
        fig.add_hline(y=data["ACTUAL_FINAL_COST"], line=dict(color="black", dash="dot"))
        st.plotly_chart(style_fig(fig, height=360, title="Distribusi Biaya"), use_container_width=True)
    with c2:
        fig = go.Figure()
        for t in range(T):
            fig.add_trace(go.Violin(y=data["Results"][t]["DaysDistArray"], x0=labels[t], name=labels[t],
                                     line_color=C["accent"], fillcolor="rgba(30,111,166,0.25)", showlegend=False, points=False))
        fig.add_hline(y=data["ACTUAL_FINAL_DAYS"], line=dict(color="black", dash="dot"))
        st.plotly_chart(style_fig(fig, height=360, title="Distribusi Waktu"), use_container_width=True)

# =========================================================================
# TAB 5: VALIDASI MODEL
# =========================================================================
with tabs[5]:
    section_header("Metrik Validasi")
    c1, c2, c3, c4 = st.columns(4)
    with c1: metric_card("MAPE Biaya", f"{v['mape_cost']:.2f}%", "EVM Tradisional")
    with c2: metric_card("MAPE Waktu", f"{v['mape_days']:.2f}%", "EVM Tradisional")
    with c3: metric_card("Cakupan Biaya", f"{v['cov_cost_pct']:.1f}%", f"{v['cov_cost_n'][0]}/{v['cov_cost_n'][1]} periode")
    with c4: metric_card("Cakupan Waktu", f"{v['cov_days_pct']:.1f}%", f"{v['cov_days_n'][0]}/{v['cov_days_n'][1]} periode")

    insight_box(
        "Cara Membaca Metrik Ini",
        "<b>MAPE</b> mengukur rata-rata persentase kesalahan estimasi EVM tradisional (titik tunggal) "
        "terhadap nilai aktual akhir - semakin kecil semakin akurat. <b>Cakupan interval</b> mengukur "
        "proporsi periode di mana nilai aktual berada dalam selang prediksi P5-P95 Augmented EVM - "
        "idealnya mendekati 90% (selang kepercayaan yang digunakan)."
    )

    section_header("Selisih Prediksi terhadap Nilai Aktual")
    err_cost_trad = np.abs(data["ACTUAL_FINAL_COST"] - df["CostTrad"])
    err_cost_aug = np.abs(data["ACTUAL_FINAL_COST"] - df["CostAugMean"])
    err_days_trad = np.abs(data["ACTUAL_FINAL_DAYS"] - df["DaysTrad"])
    err_days_aug = np.abs(data["ACTUAL_FINAL_DAYS"] - df["DaysAugMean"])
    c1, c2 = st.columns(2)
    with c1:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=x, y=err_cost_trad, name="Tradisional", line=dict(color=C["navy"], dash="dash")))
        fig.add_trace(go.Scatter(x=x, y=err_cost_aug, name="Augmented (mean)", line=dict(color=C["red"])))
        fig.update_xaxes(tickmode="array", tickvals=x, ticktext=labels)
        st.plotly_chart(style_fig(fig, height=340, title="Error Biaya"), use_container_width=True)
    with c2:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=x, y=err_days_trad, name="Tradisional", line=dict(color=C["navy"], dash="dash")))
        fig.add_trace(go.Scatter(x=x, y=err_days_aug, name="Augmented (mean)", line=dict(color=C["red"])))
        fig.update_xaxes(tickmode="array", tickvals=x, ticktext=labels)
        st.plotly_chart(style_fig(fig, height=340, title="Error Waktu"), use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.download_button("Unduh Hasil Lengkap (CSV)", df.to_csv(index=False).encode("utf-8"),
                        "hasil_augmented_evm_dashboard.csv", "text/csv")

st.markdown("---")
st.caption(
    "Risk-Based EVM Analyzer - dibangun untuk mendukung evaluasi kinerja biaya & waktu proyek EPC "
    "berbasis risiko. Metodologi: Saaty (1980) AHP, Duc (2025) Augmented EVM."
)
logo_img_html = f'<img src="data:image/png;base64,{LOGO_B64}" alt="Logo {UNIVERSITAS}" />' if LOGO_B64 else ""
st.markdown(f"""
<div class="author-card">
  {logo_img_html}
  <div>
    <div class="author-label">Disusun oleh</div>
    <div class="author-name">{PENYUSUN_NAMA}</div>
    <div class="author-nim">NIM {PENYUSUN_NIM} — {PRODI}</div>
    <div class="author-nim">{JURUSAN}, {FAKULTAS}, {UNIVERSITAS}</div>
  </div>
</div>
""", unsafe_allow_html=True)
